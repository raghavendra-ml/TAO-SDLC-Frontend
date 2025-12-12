"""
API Specification Parser
Converts Excel-based API requirements into structured API specifications
"""
import openpyxl
from typing import Dict, List, Any, Optional
import re
import json


class APISpecParser:
    """Parser for API specifications from Excel documents"""
    
    def __init__(self):
        self.supported_methods = ['GET', 'POST', 'PUT', 'DELETE', 'PATCH', 'OPTIONS', 'HEAD']
        self.common_request_bodies = {}
        self.common_responses = {}
        self.common_error_codes = {}
    
    def parse_api_excel(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file containing API specifications
        
        Expected Excel structure:
        - Tab 1: Overview (project info, base URL, authentication) - OPTIONAL
        - Tab 2+: API endpoints (one tab per endpoint group or all in one)
        
        Expected columns in API tabs (flexible, will auto-detect):
        - Endpoint/Path/URL
        - Method (GET, POST, etc.)
        - Description/Summary
        - Request Body/Parameters/Payload
        - Response Body/Response
        - Status Codes/HTTP Status
        - Authentication Required/Auth/Security
        - Tags/Category/Group
        
        Returns:
            Structured API specification data
        """
        try:
            print(f"[INFO] Loading Excel file: {file_path}")
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            print(f"[INFO] Found {len(workbook.sheetnames)} sheets: {workbook.sheetnames}")
            
            api_spec = {
                'info': {
                    'title': 'API Specification',
                    'version': '1.0.0',
                    'description': ''
                },
                'servers': [],
                'paths': {},
                'components': {
                    'schemas': {},
                    'securitySchemes': {}
                },
                'tags': []
            }
            
            # First pass: Parse common/shared elements
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if self._is_common_sheet(sheet_name):
                    print(f"[INFO] Parsing common sheet: {sheet_name}")
                    self._parse_common_sheet(sheet, sheet_name)
            
            # Second pass: Parse each sheet
            endpoints_found = 0
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Skip common sheets (already parsed)
                if self._is_common_sheet(sheet_name):
                    continue
                
                # Try to detect if it's an overview/info sheet
                if self._is_overview_sheet(sheet, sheet_name):
                    print(f"[INFO] Parsing overview sheet: {sheet_name}")
                    self._parse_overview_sheet(sheet, api_spec)
                else:
                    # Assume it's an API endpoint sheet
                    print(f"[INFO] Parsing API endpoints sheet: {sheet_name}")
                    before_count = len(api_spec['paths'])
                    self._parse_api_endpoints_sheet(sheet, sheet_name, api_spec)
                    after_count = len(api_spec['paths'])
                    endpoints_added = after_count - before_count
                    endpoints_found += endpoints_added
                    print(f"[INFO] Found {endpoints_added} endpoints in sheet '{sheet_name}'")
            
            # Third pass: Apply common elements to all endpoints
            self._apply_common_elements(api_spec)
            
            # If no servers defined, add default
            if not api_spec['servers']:
                api_spec['servers'] = [
                    {'url': 'https://api.example.com/v1', 'description': 'API Server'}
                ]
            
            print(f"[SUCCESS] Total endpoints parsed: {len(api_spec['paths'])}")
            
            if len(api_spec['paths']) == 0:
                raise Exception(
                    "No API endpoints found in Excel file. Please ensure your Excel file has:\n"
                    "1. A sheet with API endpoint data\n"
                    "2. Required columns: Endpoint/Path, Method (GET/POST/etc.), Description\n"
                    "3. At least one row of endpoint data\n\n"
                    "Example format:\n"
                    "| Endpoint | Method | Description | Request Body | Response |\n"
                    "| /api/users | GET | Get all users | - | User list |\n"
                )
            
            return api_spec
            
        except Exception as e:
            error_msg = str(e)
            print(f"[ERROR] Failed to parse API Excel file: {error_msg}")
            if "No API endpoints found" in error_msg:
                raise Exception(error_msg)
            else:
                raise Exception(
                    f"Error parsing API Excel file: {error_msg}\n\n"
                    "Please ensure your Excel file follows this format:\n"
                    "- Required columns: Endpoint/Path, Method, Description\n"
                    "- Optional columns: Request Body, Response, Status Codes, Authentication\n"
                    "- First row should be headers\n"
                    "- Data rows should follow headers"
                )
    
    def _is_common_sheet(self, sheet_name: str) -> bool:
        """Detect if sheet contains common/shared elements"""
        sheet_name_lower = sheet_name.lower()
        return any(word in sheet_name_lower for word in ['common', 'shared', 'global', 'default'])
    
    def _is_overview_sheet(self, sheet, sheet_name: str) -> bool:
        """Detect if sheet is an overview/info sheet"""
        sheet_name_lower = sheet_name.lower()
        if any(word in sheet_name_lower for word in ['overview', 'info', 'general', 'config', 'setup']):
            return True
        
        # Check first few rows for key-value pairs
        rows = list(sheet.iter_rows(values_only=True, max_row=10))
        key_value_count = 0
        for row in rows:
            if row and len(row) >= 2 and row[0] and row[1]:
                key = str(row[0]).lower()
                if any(word in key for word in ['title', 'version', 'url', 'base', 'server', 'auth']):
                    key_value_count += 1
        
        return key_value_count >= 2
    
    def _parse_common_sheet(self, sheet, sheet_name: str):
        """Parse common/shared elements like error codes and request bodies"""
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return
        
        # Try to detect what kind of common data this is
        headers = [str(cell).lower() if cell else '' for cell in rows[0]]
        
        # Check if this looks like error codes
        if any('error' in h or 'status' in h or 'code' in h for h in headers):
            self._parse_common_error_codes(rows)
        
        # Check if this looks like request/response bodies
        elif any('request' in h or 'body' in h or 'response' in h for h in headers):
            self._parse_common_bodies(rows)
        
        # Fallback: Try to parse as error codes if we find numeric codes
        else:
            self._parse_common_error_codes(rows)
    
    def _parse_common_error_codes(self, rows: List[tuple]):
        """Parse common error code definitions"""
        if not rows or len(rows) < 2:
            return
        
        headers = [str(cell).lower() if cell else '' for cell in rows[0]]
        
        # Find relevant columns
        code_col = next((i for i, h in enumerate(headers) if 'code' in h or 'status' in h), 0)
        desc_col = next((i for i, h in enumerate(headers) if 'description' in h or 'message' in h), 1 if len(headers) > 1 else 0)
        
        # Parse error codes
        for row in rows[1:]:
            if not row or not any(row):
                continue
            
            code = str(row[code_col]) if len(row) > code_col and row[code_col] else ''
            description = str(row[desc_col]) if len(row) > desc_col and row[desc_col] else ''
            
            # Extract numeric code
            code_match = re.search(r'\b([4-5]\d{2})\b', code)
            if code_match:
                status_code = code_match.group(1)
                if not description:
                    description = self._get_default_error_description(status_code)
                
                self.common_error_codes[status_code] = {
                    'description': description,
                    'content': {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'properties': {
                                    'error': {'type': 'string', 'description': 'Error message'},
                                    'code': {'type': 'string', 'description': 'Error code'},
                                    'message': {'type': 'string', 'description': 'Detailed error message'}
                                }
                            }
                        }
                    }
                }
    
    def _parse_common_bodies(self, rows: List[tuple]):
        """Parse common request/response body definitions"""
        # This method can be expanded based on specific needs
        # For now, we'll focus on common error codes
        pass
    
    def _get_default_error_description(self, status_code: str) -> str:
        """Get default description for HTTP status code"""
        default_descriptions = {
            '400': 'Bad Request - Invalid parameters provided',
            '401': 'Unauthorized - Authentication credentials missing or invalid',
            '403': 'Forbidden - Access denied',
            '404': 'Not Found - Resource not found',
            '409': 'Conflict - Resource conflict',
            '422': 'Unprocessable Entity - Validation failed',
            '429': 'Too Many Requests - Rate limit exceeded',
            '500': 'Internal Server Error - An unexpected error occurred',
            '502': 'Bad Gateway - Invalid response from upstream server',
            '503': 'Service Unavailable - Service temporarily unavailable'
        }
        return default_descriptions.get(status_code, f'HTTP {status_code} Error')
    
    def _apply_common_elements(self, api_spec: Dict[str, Any]):
        """Apply common error codes and elements to all endpoints"""
        if not self.common_error_codes:
            # If no common sheet found, use default error codes
            self.common_error_codes = {
                '400': {
                    'description': 'Bad Request - Invalid parameters provided',
                    'content': {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'properties': {
                                    'error': {'type': 'string'},
                                    'message': {'type': 'string'}
                                }
                            }
                        }
                    }
                },
                '401': {
                    'description': 'Unauthorized - Authentication credentials missing or invalid',
                    'content': {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'properties': {
                                    'error': {'type': 'string'},
                                    'message': {'type': 'string'}
                                }
                            }
                        }
                    }
                },
                '404': {
                    'description': 'Not Found - Resource not found',
                    'content': {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'properties': {
                                    'error': {'type': 'string'},
                                    'message': {'type': 'string'}
                                }
                            }
                        }
                    }
                },
                '500': {
                    'description': 'Internal Server Error - An unexpected error occurred',
                    'content': {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'properties': {
                                    'error': {'type': 'string'},
                                    'message': {'type': 'string'}
                                }
                            }
                        }
                    }
                }
            }
        
        # Apply common error codes to all endpoints
        for path, methods in api_spec['paths'].items():
            for method, operation in methods.items():
                if method.lower() in ['get', 'post', 'put', 'delete', 'patch']:
                    if 'responses' not in operation:
                        operation['responses'] = {}
                    
                    # Add common error codes if not already present
                    for code, response_spec in self.common_error_codes.items():
                        if code not in operation['responses']:
                            operation['responses'][code] = response_spec.copy()
    
    def _parse_overview_sheet(self, sheet, api_spec: Dict[str, Any]):
        """Parse overview/info sheet for general API information"""
        rows = list(sheet.iter_rows(values_only=True))
        
        for row in rows:
            if not row or not any(row):
                continue
            
            key = str(row[0]).lower().strip() if row[0] else ''
            value = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            
            if not key or not value:
                continue
            
            if 'title' in key or 'api name' in key or 'name' in key:
                api_spec['info']['title'] = value
            elif 'version' in key:
                api_spec['info']['version'] = value
            elif 'description' in key:
                api_spec['info']['description'] = value
            elif 'base url' in key or 'server' in key or 'host' in key:
                api_spec['servers'].append({
                    'url': value,
                    'description': 'API Server'
                })
            elif 'auth' in key and value:
                # Parse authentication type
                auth_type = value.lower()
                if 'bearer' in auth_type or 'jwt' in auth_type:
                    api_spec['components']['securitySchemes']['BearerAuth'] = {
                        'type': 'http',
                        'scheme': 'bearer',
                        'bearerFormat': 'JWT'
                    }
                elif 'api key' in auth_type or 'apikey' in auth_type:
                    api_spec['components']['securitySchemes']['ApiKeyAuth'] = {
                        'type': 'apiKey',
                        'in': 'header',
                        'name': 'X-API-Key'
                    }
                elif 'oauth' in auth_type:
                    api_spec['components']['securitySchemes']['OAuth2'] = {
                        'type': 'oauth2',
                        'flows': {}
                    }
    
    def _parse_api_endpoints_sheet(self, sheet, sheet_name: str, api_spec: Dict[str, Any]):
        """Parse API endpoints from a sheet"""
        try:
            rows = list(sheet.iter_rows(values_only=True))
            
            if not rows or len(rows) < 2:
                print(f"[WARNING] Sheet '{sheet_name}' has insufficient rows (need at least 2)")
                return
            
            # Find header row (first non-empty row)
            header_row_idx = 0
            for i, row in enumerate(rows):
                if any(row):
                    header_row_idx = i
                    break
            
            header_row = rows[header_row_idx]
            headers = [str(h).lower().strip() if h else '' for h in header_row]
            
            print(f"[DEBUG] Headers in sheet '{sheet_name}': {headers}")
            
            # Map column indices
            col_map = self._create_column_map(headers)
            
            print(f"[DEBUG] Column mapping for '{sheet_name}': {col_map}")
            
            # Check if we found essential columns
            if 'path' not in col_map:
                print(f"[WARNING] No 'Endpoint/Path' column found in sheet '{sheet_name}'. Skipping.")
                print(f"[INFO] Available headers: {headers}")
                print(f"[INFO] Please include a column with one of these names: endpoint, path, url, api, route, resource")
                print(f"[INFO] First 3 data rows for reference:")
                for i, row in enumerate(rows[header_row_idx + 1:header_row_idx + 4]):
                    print(f"[INFO]   Row {i+1}: {row}")
                return
            
            # Add sheet as a tag if it has meaningful name
            tag_name = sheet_name.replace('_', ' ').replace('-', ' ').title()
            if tag_name and tag_name not in [t['name'] for t in api_spec['tags']]:
                api_spec['tags'].append({
                    'name': tag_name,
                    'description': f'{tag_name} endpoints'
                })
            
            # Parse each endpoint row
            endpoints_parsed = 0
            for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
                if not row or not any(row):
                    continue
                
                try:
                    endpoint = self._parse_endpoint_row(row, col_map, tag_name)
                    if endpoint and endpoint['path']:
                        path = endpoint['path']
                        method = endpoint['method'].lower()
                        
                        # Initialize path if not exists
                        if path not in api_spec['paths']:
                            api_spec['paths'][path] = {}
                        
                        # Add method to path
                        api_spec['paths'][path][method] = endpoint['spec']
                        endpoints_parsed += 1
                except Exception as row_error:
                    print(f"[WARNING] Error parsing row {row_idx} in sheet '{sheet_name}': {str(row_error)}")
                    continue
            
            print(f"[INFO] Successfully parsed {endpoints_parsed} endpoints from sheet '{sheet_name}'")
            
        except Exception as sheet_error:
            print(f"[ERROR] Error processing sheet '{sheet_name}': {str(sheet_error)}")
            import traceback
            traceback.print_exc()
    
    def _create_column_map(self, headers: List[str]) -> Dict[str, int]:
        """Create a mapping of column names to indices"""
        col_map = {}
        
        print(f"[DEBUG] Creating column map from headers: {headers}")
        
        for i, header in enumerate(headers):
            if not header:
                continue
            
            # Convert to lowercase and remove extra spaces for case-insensitive matching
            header_lower = str(header).lower().strip()
            
            print(f"[DEBUG] Processing header[{i}]: '{header}' (lowercase: '{header_lower}')")
            
            # Endpoint/Path - be more flexible
            if any(word in header_lower for word in ['endpoint', 'path', 'url', 'api', 'route', 'resource', 'operationid']):
                if 'path' not in col_map:
                    col_map['path'] = i
                    print(f"[DEBUG] Mapped 'path' to column {i}: '{header}'")
            
            # Method - be more flexible
            elif any(word in header_lower for word in ['method', 'http', 'verb', 'type', 'action']):
                col_map['method'] = i
                print(f"[DEBUG] Mapped 'method' to column {i}: '{header}'")
            
            # Description - be more flexible
            elif any(word in header_lower for word in ['description', 'summary', 'detail', 'purpose', 'function', 'info']):
                if 'description' not in col_map:
                    col_map['description'] = i
                    print(f"[DEBUG] Mapped 'description' to column {i}: '{header}'")
            
            # Request Body
            elif 'request' in header_lower and ('body' in header_lower or 'payload' in header_lower or 'data' in header_lower or 'schema' in header_lower):
                col_map['request_body'] = i
                print(f"[DEBUG] Mapped 'request_body' to column {i}: '{header}'")
            
            # Request Parameters
            elif ('request' in header_lower or 'input' in header_lower) and ('param' in header_lower or 'query' in header_lower):
                col_map['request_params'] = i
                print(f"[DEBUG] Mapped 'request_params' to column {i}: '{header}'")
            elif any(word in header_lower for word in ['parameter', 'params', 'query']):
                if 'request_params' not in col_map:
                    col_map['request_params'] = i
                    print(f"[DEBUG] Mapped 'request_params' to column {i}: '{header}'")
            
            # Response
            elif 'response' in header_lower:
                if 'example' in header_lower:
                    col_map['response_example'] = i
                    print(f"[DEBUG] Mapped 'response_example' to column {i}: '{header}'")
                else:
                    col_map['response'] = i
                    print(f"[DEBUG] Mapped 'response' to column {i}: '{header}'")
            
            # Status Codes
            elif any(word in header_lower for word in ['status', 'code', 'http code']):
                if 'status_codes' not in col_map:
                    col_map['status_codes'] = i
                    print(f"[DEBUG] Mapped 'status_codes' to column {i}: '{header}'")
            
            # Authentication
            elif any(word in header_lower for word in ['auth', 'security', 'authentication', 'authorization']):
                col_map['auth'] = i
                print(f"[DEBUG] Mapped 'auth' to column {i}: '{header}'")
            
            # Tags/Category
            elif any(word in header_lower for word in ['tag', 'category', 'group', 'module']):
                col_map['tags'] = i
                print(f"[DEBUG] Mapped 'tags' to column {i}: '{header}'")
        
        print(f"[DEBUG] Final column map: {col_map}")
        return col_map
    
    def _parse_endpoint_row(self, row: tuple, col_map: Dict[str, int], default_tag: str) -> Optional[Dict[str, Any]]:
        """Parse a single endpoint row"""
        try:
            path = self._get_cell_value(row, col_map.get('path'))
            method = self._get_cell_value(row, col_map.get('method'), 'GET')
            
            if not path or path == 'none':
                return None
            
            # Clean path
            path = path.strip()
            if not path.startswith('/'):
                path = '/' + path
            
            # Ensure method is valid
            method = method.upper().strip()
            if method not in self.supported_methods:
                # Try to extract method from path or default to GET
                method = 'GET'
            
            description = self._get_cell_value(row, col_map.get('description'), '')
            request_body = self._get_cell_value(row, col_map.get('request_body'), '')
            request_params = self._get_cell_value(row, col_map.get('request_params'), '')
            response = self._get_cell_value(row, col_map.get('response'), '')
            response_example = self._get_cell_value(row, col_map.get('response_example'), '')
            status_codes = self._get_cell_value(row, col_map.get('status_codes'), '200')
            auth_required = self._get_cell_value(row, col_map.get('auth'), '')
            tags = self._get_cell_value(row, col_map.get('tags'), default_tag)
            
            print(f"[DEBUG] Parsing endpoint {method} {path}")
            print(f"[DEBUG] Response body from Excel: {response[:200] if response else 'EMPTY'}")
            
            # Build OpenAPI operation object
            operation = {
                'tags': [tags] if isinstance(tags, str) else tags,
                'summary': description or f'{method} {path}',
                'description': description,
                'responses': self._parse_responses(response, status_codes, response_example)
            }
            
            # Add parameters if present
            params = self._parse_parameters(path, request_params)
            if params:
                operation['parameters'] = params
            
            print(f"[DEBUG] Parameters for {method} {path}: {len(params)} params")
            if params:
                print(f"[DEBUG] Parameter details: {[p['name'] for p in params]}")
            
            # Add request body if present (for POST, PUT, PATCH)
            if method in ['POST', 'PUT', 'PATCH'] and request_body:
                operation['requestBody'] = self._parse_request_body(request_body, '')
                print(f"[DEBUG] Request body added for {method} {path}")
            elif method in ['POST', 'PUT', 'PATCH']:
                print(f"[DEBUG] No request body for {method} {path} (request_body: '{request_body}')")
            
            # Add security if required
            if auth_required and str(auth_required).lower() in ['yes', 'true', 'required', 'y']:
                operation['security'] = [{'BearerAuth': []}]
            
            return {
                'path': path,
                'method': method,
                'spec': operation
            }
            
        except Exception as e:
            print(f"Error parsing endpoint row: {str(e)}")
            return None
    
    def _get_cell_value(self, row: tuple, col_index: Optional[int], default: str = '') -> str:
        """Safely get cell value"""
        if col_index is None or col_index >= len(row):
            return default
        
        value = row[col_index]
        if value is None:
            return default
        
        return str(value).strip()
    
    def _parse_parameters(self, path: str, params_str: str) -> List[Dict[str, Any]]:
        """Parse path and query parameters"""
        parameters = []
        
        # Extract path parameters from path (e.g., /users/{id})
        path_params = re.findall(r'\{(\w+)\}', path)
        for param in path_params:
            parameters.append({
                'name': param,
                'in': 'path',
                'required': True,
                'schema': {'type': 'string'},
                'description': f'Path parameter: {param}'
            })
        
        # Parse additional parameters from params_str
        if params_str and params_str.lower() != 'none':
            # Support multiple formats
            # Format 1: "id (string), name (string), limit (integer)"
            # Format 2: "id: string, name: string"
            # Format 3: "id, name, limit"
            
            # Try format with types in parentheses
            param_pattern = r'(\w+)\s*\((\w+)\)'
            matches = re.finditer(param_pattern, params_str)
            
            found_params = False
            for match in matches:
                found_params = True
                param_name = match.group(1).strip()
                param_type = match.group(2).strip().lower()
                
                # Skip if already added as path param
                if any(p['name'] == param_name for p in parameters):
                    continue
                
                # Map types
                type_map = {
                    'int': 'integer', 'integer': 'integer', 'number': 'integer',
                    'string': 'string', 'str': 'string', 'text': 'string',
                    'bool': 'boolean', 'boolean': 'boolean',
                    'float': 'number', 'double': 'number', 'decimal': 'number',
                    'array': 'array', 'list': 'array',
                    'object': 'object', 'dict': 'object'
                }
                
                parameters.append({
                    'name': param_name,
                    'in': 'query',
                    'required': False,
                    'schema': {'type': type_map.get(param_type, 'string')},
                    'description': f'Query parameter: {param_name}'
                })
            
            # If no matches, try simple comma-separated names
            if not found_params:
                param_names = [p.strip() for p in params_str.split(',') if p.strip()]
                for param_name in param_names:
                    if param_name and not any(p['name'] == param_name for p in parameters):
                        parameters.append({
                            'name': param_name,
                            'in': 'query',
                            'required': False,
                            'schema': {'type': 'string'},
                            'description': f'Query parameter: {param_name}'
                        })
        
        return parameters
    
    def _parse_request_body(self, body_str: str, example_str: str = '') -> Dict[str, Any]:
        """Parse request body schema"""
        if not body_str or body_str.lower() == 'none':
            return {
                'content': {
                    'application/json': {
                        'schema': {'type': 'object'}
                    }
                }
            }
        
        try:
            # Try to parse as JSON
            if body_str.strip().startswith('{'):
                body_obj = json.loads(body_str)
                schema = self._infer_schema_from_object(body_obj)
                example = body_obj
            else:
                # Parse text description into schema
                schema = self._parse_schema_from_text(body_str)
                example = json.loads(example_str) if example_str.strip().startswith('{') else None
            
            request_body = {
                'required': True,
                'content': {
                    'application/json': {
                        'schema': schema
                    }
                }
            }
            
            if example:
                request_body['content']['application/json']['example'] = example
            
            return request_body
            
        except Exception as e:
            # Fallback to generic object
            return {
                'required': True,
                'content': {
                    'application/json': {
                        'schema': {
                            'type': 'object',
                            'description': body_str
                        }
                    }
                }
            }
    
    def _parse_responses(self, response_str: str, status_codes: str, example_str: str = '') -> Dict[str, Any]:
        """Parse response definitions"""
        responses = {}
        
        # Parse status codes
        if status_codes:
            codes = [c.strip() for c in str(status_codes).split(',')]
        else:
            codes = ['200']
        
        for code in codes:
            # Extract just the number
            code_num = re.search(r'\d+', code)
            if not code_num:
                continue
            
            code = code_num.group()
            code_int = int(code)
            
            # Default descriptions
            desc_map = {
                200: 'Successful response',
                201: 'Created successfully',
                204: 'No content',
                400: 'Bad request',
                401: 'Unauthorized',
                403: 'Forbidden',
                404: 'Not found',
                500: 'Internal server error'
            }
            
            response_def = {
                'description': desc_map.get(code_int, f'Response {code}')
            }
            
            # Add schema for success responses
            if 200 <= code_int < 300 and response_str and response_str.lower() != 'none':
                try:
                    # Try to parse as JSON
                    if response_str.strip().startswith('{') or response_str.strip().startswith('['):
                        response_obj = json.loads(response_str)
                        schema = self._infer_schema_from_object(response_obj)
                        example = response_obj
                    else:
                        # Parse from text description
                        schema = self._parse_schema_from_text(response_str)
                        # Try to parse example if provided
                        example = None
                        if example_str and (example_str.strip().startswith('{') or example_str.strip().startswith('[')):
                            try:
                                example = json.loads(example_str)
                            except:
                                pass
                    
                    print(f"[DEBUG] Parsed response schema for {code}: {schema}")
                    
                    response_def['content'] = {
                        'application/json': {
                            'schema': schema
                        }
                    }
                    
                    if example:
                        response_def['content']['application/json']['example'] = example
                        
                except Exception as e:
                    print(f"[WARNING] Failed to parse response schema for {code}: {str(e)}")
                    print(f"[DEBUG] Response string: {response_str[:200]}")
                    # Fallback to simple schema with description
                    response_def['content'] = {
                        'application/json': {
                            'schema': {
                                'type': 'object',
                                'description': response_str
                            }
                        }
                    }
            
            responses[str(code)] = response_def
        
        # Ensure at least one response
        if not responses:
            responses['200'] = {
                'description': 'Successful response'
            }
        
        return responses
    
    def _parse_schema_from_text(self, text: str) -> Dict[str, Any]:
        """Parse schema from text description"""
        schema = {
            'type': 'object',
            'properties': {}
        }
        
        if not text or text.lower() == 'none':
            return schema
        
        # Parse field definitions
        # Format 1: "field1 (type): description, field2 (type)"
        field_pattern = r'(\w+)\s*\((\w+)\)'
        matches = re.finditer(field_pattern, text)
        
        found_fields = False
        for match in matches:
            found_fields = True
            field_name = match.group(1).strip()
            field_type = match.group(2).strip().lower()
            
            # Map types
            type_map = {
                'int': 'integer', 'integer': 'integer', 'number': 'integer',
                'string': 'string', 'str': 'string', 'text': 'string',
                'bool': 'boolean', 'boolean': 'boolean',
                'float': 'number', 'double': 'number',
                'array': 'array', 'list': 'array',
                'object': 'object', 'dict': 'object'
            }
            
            schema['properties'][field_name] = {
                'type': type_map.get(field_type, 'string')
            }
        
        # Format 2: Try to parse "field1: value, field2: value" format
        if not found_fields and ':' in text:
            # Split by newline or comma
            parts = re.split(r'[,\n]', text)
            for part in parts:
                if ':' in part:
                    field_parts = part.split(':', 1)
                    if len(field_parts) == 2:
                        field_name = field_parts[0].strip()
                        field_value = field_parts[1].strip()
                        
                        # Clean field name
                        field_name = re.sub(r'[^\w]', '', field_name)
                        if field_name:
                            found_fields = True
                            # Infer type from value
                            if field_value.lower() in ['true', 'false']:
                                field_type = 'boolean'
                            elif field_value.isdigit():
                                field_type = 'integer'
                            elif re.match(r'^\d+\.\d+$', field_value):
                                field_type = 'number'
                            else:
                                field_type = 'string'
                            
                            schema['properties'][field_name] = {'type': field_type}
        
        # If no structured fields found, try simple comma-separated fields
        if not found_fields and ',' in text:
            fields = [f.strip() for f in text.split(',') if f.strip()]
            for field in fields:
                # Remove special characters
                field_clean = re.sub(r'[^\w\s]', '', field).strip()
                if field_clean and len(field_clean) > 0:
                    schema['properties'][field_clean] = {'type': 'string', 'description': field}
                    found_fields = True
        
        # Format 3: Try to extract words that look like field names
        if not found_fields:
            # Look for camelCase or snake_case identifiers
            potential_fields = re.findall(r'\b[a-z_][a-zA-Z0-9_]*\b', text)
            if potential_fields:
                for field in potential_fields[:5]:  # Limit to first 5 to avoid noise
                    if len(field) > 2:  # Ignore very short words
                        schema['properties'][field] = {'type': 'string'}
                        found_fields = True
        
        print(f"[DEBUG] Schema from text: {schema}")
        print(f"[DEBUG] Original text: {text[:200]}")
        
        return schema if schema['properties'] else {'type': 'object', 'description': text}
    
    def _infer_schema_from_object(self, obj: Any) -> Dict[str, Any]:
        """Infer OpenAPI schema from a JSON object"""
        if isinstance(obj, dict):
            schema = {
                'type': 'object',
                'properties': {}
            }
            
            for key, value in obj.items():
                schema['properties'][key] = self._infer_schema_from_object(value)
            
            return schema
            
        elif isinstance(obj, list):
            if obj:
                return {
                    'type': 'array',
                    'items': self._infer_schema_from_object(obj[0])
                }
            else:
                return {
                    'type': 'array',
                    'items': {'type': 'object'}
                }
        
        elif isinstance(obj, bool):
            return {'type': 'boolean'}
        
        elif isinstance(obj, int):
            return {'type': 'integer'}
        
        elif isinstance(obj, float):
            return {'type': 'number'}
        
        elif isinstance(obj, str):
            return {'type': 'string'}
        
        else:
            return {'type': 'string'}

